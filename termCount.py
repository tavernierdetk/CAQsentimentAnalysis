import json
import openpyxl as xl
import pickle



def termCount(textArray):
    dictionary = {}

    for text in textArray:
        tokens = text.split()
        for i in tokens:
            if i not in dictionary:
                dictionary[i] = 1
            else:
                dictionary[i] +=1
    return dictionary

def output(dictionary, outputFilepath):

    wb = xl.Workbook()
    ws = wb.active

    count = 1

    for i in dictionary.keys():
        ws.cell(row=count, column=1).value = i
        ws.cell(row=count, column=2).value = dictionary[i]
        count += 1

    wb.save(outputFilepath)




if __name__ == "__main__":
    with open('JSONFiles/envVars.json') as file:
        envVars = json.load(file)
        file.close()
    filepath = envVars["frenchLexFilepath"]
    with open('JSONFiles/caqai-414e9-FB_POSTS-export.json') as file:
        dataset = json.load(file)
        file.close()

    all_posts = []
    all_comments = []

    counter = 0
    for post in dataset:
        if 'comments' not in post:
            continue
        if 'message' not in post:
            continue
        all_posts.append(post['message'])
        counter +=1

        for comment in post['comments']:
            all_comments.append(comment['message'])

    with open("pickleFiles/comments_corpus.pkl", 'wb') as file:
        pickle.dump(all_comments, file)
        file.close()

    with open("pickleFiles/posts_corpus.pkl", 'wb') as file:
        pickle.dump(all_posts, file)
        file.close()


