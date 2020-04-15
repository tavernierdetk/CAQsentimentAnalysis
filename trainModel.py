import pickle
import pandas as pd
import openpyxl as xl
from sklearn.ensemble import RandomForestClassifier
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score
from sklearn.metrics import precision_recall_curve
from sklearn.metrics import average_precision_score
from sklearn.metrics import auc
from sklearn.metrics import f1_score
from sklearn.metrics import recall_score as rScore
from sklearn.metrics import roc_auc_score
from sklearn.metrics import log_loss as logloss
from sklearn.metrics import brier_score_loss

def runTest(model, X_test, y_test, X):

    feat_importances = pd.Series(model.feature_importances_, index=X.columns)
    for i in feat_importances.items():
        print(i)

    wb = xl.Workbook()
    ws = wb.active
    predictions_bool= model.predict(X_test)
    predictions = model.predict_proba (X_test)
    predictions_prob = predictions[:, 1]

    accuracy = accuracy_score(y_test, predictions_bool)
    roc_auc = roc_auc_score(y_test, predictions_prob)
    log_loss = logloss(y_test, predictions_prob)
    f1Score = f1_score(y_test, predictions_bool, average='weighted')
    precision, recall, thresholds = precision_recall_curve(y_test, predictions_prob)
    recall_score = rScore(y_test,predictions_bool, average='weighted')
    brierScore = brier_score_loss(y_test, predictions_prob)

    avgPrecision = average_precision_score(y_test, predictions_bool, average='weighted')
    auc_score = auc(recall, precision)


    metrics = {'accuracy': accuracy,
               'roc_auc': roc_auc,
               'neg_log_loss': log_loss,
               'f1': f1Score,
               'average_precision': avgPrecision,
               'auc_score': auc_score,
               'recall_score': recall_score,
               'brierScore': brierScore
               }

    for index, i in enumerate(metrics):
        ws.cell(row=index+2,column=1).value = i + ": "
        ws.cell(row=index+2,column=2).value = metrics[i]
    wb.save('model_score.xlsx')

def trainModel(df):
    model = RandomForestClassifier(random_state=0, n_estimators=200, verbose=0)

    X = df.drop(columns=["is_popular", "avg_comment_sentiment","number_of_shares", 'day_posted', 'number_of_comments','message'])

    y = df["number_of_shares"]

    # Define that there will be X and y for training and testing
    X_train, X_test, y_train, y_test = train_test_split (X, y, test_size = 0.2)
    model.fit(X_train, y_train)
    runTest(model, X_test, y_test, X)

    with open('pickleFiles/model.pkl', 'wb') as file:
        pickle.dump(model, file)

    return X, model

if __name__ == "__main__":
    print('go')

    with open('pickleFiles/outputWithSimilarity.pkl', 'rb') as file:
        df = pickle.load(file)

    X, model = trainModel(df)