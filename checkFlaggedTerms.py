import json
import openpyxl as xl

def returnFlaggedTerms(text, dictionary):
    tokens = text.split()
    flagged_words = []
    for token in tokens:
        if token in dictionary:
            flagged_words.append((token,dictionary[token]))
    return flagged_words

def makeDict(filepath):
    with open(filepath, encoding='utf-8') as f:
        lexicon_full_filepath = f.read()
    lex_dict = {}
    for line in lexicon_full_filepath.split('\n'):
        (word, measure) = line.strip().split('\t')[0:2]
        lex_dict[word] = float(measure)
    return lex_dict

def createExcelDict(dataset, dictionary):
    wb = xl.Workbook()
    ws = wb.active

    count = 1
    for post in dataset:

        if "message" in post:
            flagged_words = returnFlaggedTerms(post["message"],dictionary)
            ws.cell(column=1, row=count).value = post['message']
            output_string = ''
            for i in flagged_words:
                output_string += i[0]+': '+ str(i[1]) +'\n'
            ws.cell(column=2, row=count).value = output_string
            count += 1

    wb.save('outputExcel.xlsx')

if __name__ == '__main__':
    with open('JSONFiles/envVars.json') as file:
        envVars = json.load(file)
        file.close()
    filepath = envVars["frenchLexFilepath"]
    with open('JSONFiles/caqai-414e9-FB_POSTS-export.json') as file:
        dataset = json.load(file)
        file.close()
    dictionary = makeDict(filepath)
    createExcelDict(dataset, dictionary)